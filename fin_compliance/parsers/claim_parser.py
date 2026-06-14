import csv
import json
from pathlib import Path
from typing import Any, Dict, List

from fin_compliance.domain.schemas import ReimbursementClaim


class ClaimParser:
    def parse(self, path: str | Path) -> ReimbursementClaim:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix == ".json":
            return self._parse_json(path)
        if suffix in {".csv", ".tsv"}:
            return self._parse_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._parse_xlsx(path)
        raise ValueError(f"Unsupported claim file type: {path.suffix}")

    def _parse_json(self, path: Path) -> ReimbursementClaim:
        with path.open("r", encoding="utf-8") as file:
            return ReimbursementClaim(**json.load(file))

    def _parse_csv(self, path: Path) -> ReimbursementClaim:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = list(csv.DictReader(file, delimiter=delimiter))
        if not rows:
            raise ValueError(f"Empty claim file: {path}")
        return self._rows_to_claim(rows, path.stem)

    def _parse_xlsx(self, path: Path) -> ReimbursementClaim:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ImportError("Install openpyxl to parse Excel claims: pip install openpyxl") from error

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if len(raw_rows) < 2:
            raise ValueError(f"Excel claim file must include header and data rows: {path}")

        headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
        rows = []
        for raw_row in raw_rows[1:]:
            row = {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
            if any(value for value in row.values()):
                rows.append(row)
        return self._rows_to_claim(rows, path.stem)

    def _rows_to_claim(self, rows: List[Dict[str, Any]], default_claim_id: str) -> ReimbursementClaim:
        first = rows[0]
        claim: Dict[str, Any] = {
            "claim_id": first.get("claim_id") or default_claim_id,
            "employee_name": first.get("employee_name", "unknown"),
            "employee_level": first.get("employee_level", "staff"),
            "department": first.get("department", "unknown"),
            "company_name": first.get("company_name", ""),
            "trip_city": first.get("trip_city") or first.get("city", ""),
            "trip_start": first.get("trip_start", ""),
            "trip_end": first.get("trip_end", ""),
            "approval_chain": self._split_pipe(first.get("approval_chain", "")),
            "approval_files": self._split_pipe(first.get("approval_files", "")),
            "attachments": self._split_pipe(first.get("attachments", "")),
            "items": [],
        }
        for index, row in enumerate(rows, start=1):
            claim["items"].append(
                {
                    "item_id": row.get("item_id") or f"ITEM-{index:03d}",
                    "item_type": row.get("item_type", "other"),
                    "amount": float(row.get("amount", 0) or 0),
                    "currency": row.get("currency", "CNY"),
                    "city": row.get("city") or claim["trip_city"],
                    "date": row.get("date") or None,
                    "nights": int(float(row.get("nights", 1) or 1)),
                    "invoice_type": row.get("invoice_type") or None,
                    "buyer_name": row.get("buyer_name") or None,
                    "seller_name": row.get("seller_name") or None,
                    "transport_class": row.get("transport_class") or None,
                    "description": row.get("description", ""),
                }
            )
        return ReimbursementClaim(**claim)

    def _split_pipe(self, value: str) -> List[str]:
        return [item.strip() for item in str(value).split("|") if item.strip()]
